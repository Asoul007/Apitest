# -*- coding: utf-8 -*-
# @Time    : 2022/9/4 18:38
# @Author  : Asoul
import openpyxl

class HandleXls:
    """初始化"""
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read_excel(self):
        """读数据"""
        # 第一步：将excel文件加载到一个工作簿对象中
        wb = openpyxl.load_workbook(self.filename)
        # 第二步：选择文件中的表单
        sh = wb[self.sheetname]
        rows = list(sh.rows)
        # print(res,type(res))
        # 遍历第一行所有的单元格，将格子中的值添加到列表title中
        title = [i.value for i in rows[0]]
        # title = []
        # for i in res[0]:
        #     title.append(i.value)
        # print(title)

        # 遍历除第一行之外所有的行
        case_data = []
        for row in rows[1:]:
            data = [i.value for i in row]
            case_data.append(dict(zip(title, data)))
        return case_data
        # case_data = []
        # for row in res[1:]:
        #     data = []
        #     for c in row:
        #         data.append(c.value)
        #     case = dict(zip(title, data))
        #     case_data.append(case)
        # return case_data

    def write_excel(self,row,column,value):
        """写数据"""
        #新建工作簿
        wb = openpyxl.load_workbook(self.filename)
        # 新建工作表
        sh = wb[self.sheetname]
        sh.cell(row=row,column=column,value=value)
        wb.save(self.filename)

if __name__ == '__main__':
    handxls=HandleXls('cases.xlsx','register')
    print(handxls.read_excel())
    handxls.write_excel(7,7,"777")



